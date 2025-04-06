import sys
import json
import io
import os
import subprocess
import numpy as np
from scipy.interpolate import CubicSpline
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import datetime
import platform

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def calculate_forecast(data, years_to_predict, model_type):
    model_type = model_type.lower()
    years = np.array([entry['year'].year if isinstance(entry['year'], datetime.date) else entry['year'] for entry in data])
    sales = np.array([entry['totalSales'] for entry in data])

    if model_type == "cubic_spline":
        spline = CubicSpline(years, sales)
        forecast = [{"year": year_dict['year'], "totalSales": round(spline(year_dict['year']).item(), 2)}
                    for year_dict in years_to_predict]
    elif model_type == "linear":
        coeffs = np.polyfit(years, sales, 1)
        forecast = [{"year": year_dict['year'], "totalSales": round(np.polyval(coeffs, year_dict['year']), 2)}
                    for year_dict in years_to_predict]
    elif model_type == "polynomial":
        coeffs = np.polyfit(years, sales, 2)
        forecast = [{"year": year_dict['year'], "totalSales": round(np.polyval(coeffs, year_dict['year']), 2)}
                    for year_dict in years_to_predict]
    elif model_type == "exponential":
        log_sales = np.log(sales)
        coeffs = np.polyfit(years, log_sales, 1)
        a = np.exp(coeffs[1])
        b = coeffs[0]
        forecast = [{"year": year_dict['year'], "totalSales": round(a * np.exp(b * year_dict['year']), 2)}
                    for year_dict in years_to_predict]
    else:
        raise ValueError(f"Modelul '{model_type}' nu este suportat.")

    return forecast

def connect_predictions(data, forecast):
    if data:
        last_data_point = data[-1]
        forecast.insert(0, {"year": last_data_point['year'], "totalSales": last_data_point['totalSales']})
    return forecast

def create_excel_with_chart(data, forecast, output_path, model_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prediction Data"

    # Adăugare date în tabel
    ws.append(["Year", "Total Sales", "Prediction"])
    for entry in data:
        ws.append([entry['year'], entry['totalSales'], None])
    for entry in forecast:
        ws.append([entry['year'], None, entry['totalSales']])

    # Creare grafic 1: Previziuni
    chart1 = LineChart()
    chart1.title = f"Sales Prediction ({model_type.capitalize()})"
    chart1.x_axis.title = "Year"
    chart1.y_axis.title = "Total Sales"

    data_range = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(data) + len(forecast) + 1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(data) + len(forecast) + 1)
    chart1.add_data(data_range, titles_from_data=True)
    chart1.set_categories(labels)

    chart1.series[1].graphicalProperties.line.dashStyle = "sysDot"
    chart1.series[1].graphicalProperties.line.solidFill = "FF0000"
    chart1.series[0].graphicalProperties.line.solidFill = "0000FF"

    ws.add_chart(chart1, "H5")

    # Combinație date istorice + previziuni pentru calculul trend-ului
    combined_years = np.array([entry['year'] for entry in data + forecast])
    combined_sales = np.array([entry['totalSales'] for entry in data] + 
                               [entry['totalSales'] for entry in forecast if entry['totalSales'] is not None])

    # Calculare trend pe baza datelor combinate
    coeffs = np.polyfit(combined_years, combined_sales, 2)
    trend_sales = np.polyval(coeffs, combined_years)

    # Generare formulă în format text
    formula = f"y = {coeffs[0]:.4f}x² + {coeffs[1]:.4f}x + {coeffs[2]:.4f}"
    print(f"Formula trendului: {formula}")

    # Calcularea intervalelor de încredere
    residuals = combined_sales - trend_sales
    std_dev = np.std(residuals)
    trend_upper = trend_sales + 2 * std_dev
    trend_lower = trend_sales - 2 * std_dev

    # Adăugare date trend și intervale în foaia Excel
    ws.cell(row=1, column=4, value="Trend Sales")
    ws.cell(row=1, column=5, value="Trend Upper")
    ws.cell(row=1, column=6, value="Trend Lower")
    for i, (trend, upper, lower) in enumerate(zip(trend_sales, trend_upper, trend_lower), start=2):
        ws.cell(row=i, column=4, value=round(trend, 2))
        ws.cell(row=i, column=5, value=round(upper, 2))
        ws.cell(row=i, column=6, value=round(lower, 2))

    # Adăugare formulă în foaia Excel
    ws.cell(row=16, column=1, value="Trend Formula:")
    ws.cell(row=16, column=2, value=formula)

    # Creare grafic 2: Trend cu intervale
    chart2 = LineChart()
    chart2.title = "Trend of Sales (Including Forecast)"
    chart2.x_axis.title = "Year"
    chart2.y_axis.title = "Sales"

    trend_range = Reference(ws, min_col=4, min_row=1, max_row=len(data) + len(forecast) + 1)
    trend_upper_range = Reference(ws, min_col=5, min_row=1, max_row=len(data) + len(forecast) + 1)
    trend_lower_range = Reference(ws, min_col=6, min_row=1, max_row=len(data) + len(forecast) + 1)
    trend_labels = Reference(ws, min_col=1, min_row=2, max_row=len(data) + len(forecast) + 1)

    chart2.add_data(trend_range, titles_from_data=True)
    chart2.add_data(trend_upper_range, titles_from_data=True)
    chart2.add_data(trend_lower_range, titles_from_data=True)
    chart2.set_categories(trend_labels)

    chart2.series[0].graphicalProperties.line.solidFill = "0000FF"
    chart2.series[1].graphicalProperties.line.solidFill = "00FF00"
    chart2.series[2].graphicalProperties.line.solidFill = "FF0000"

    ws.add_chart(chart2, "R5")

    # Salvare fișier Excel
    wb.save(output_path)
    open_excel_file(output_path)

def open_excel_file(file_path):
    try:
        system_platform = platform.system()
        if system_platform == "Windows":
            subprocess.run(["start", file_path], shell=True, check=True)
        elif system_platform == "Darwin":  # macOS
            subprocess.run(["open", file_path], check=True)
        else:  # Linux
            subprocess.run(["xdg-open", file_path], check=True)
    except Exception as e:
        print(f"Eroare la deschiderea fișierului Excel: {e}")

if __name__ == "__main__":
    print("Argumentele primite de scriptul Python:")
    for i, arg in enumerate(sys.argv):
        print(f"Argument {i}: {arg}")
    
    input_data = json.loads(sys.argv[1])
    years_to_predict = json.loads(sys.argv[2])
    model_type = sys.argv[3]
    output_file = sys.argv[4]

    try:
        forecast = calculate_forecast(input_data, years_to_predict, model_type)
        forecast = connect_predictions(input_data, forecast)
        create_excel_with_chart(input_data, forecast, output_file, model_type)
        print(f"Fișier generat și deschis: {output_file}")
    except Exception as e:
        print(f"Eroare: {e}")
